from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SheetSummary:
    name: str
    max_row: int
    max_col: int
    header_row: int | None
    headers: list[str]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ")


def _looks_like_header_row(values: list[str]) -> bool:
    joined = " ".join(values)
    keywords = ["שם", "משפחה", "פרטי", "תלמיד", "כיתה", "תז", "ת.ז", "תעודת", "זהות"]
    hits = sum(1 for k in keywords if k in joined)
    non_empty = sum(1 for v in values if v)
    return hits >= 2 and non_empty >= 2


def _find_header_row(ws, scan_rows: int = 20, scan_cols: int = 20) -> tuple[int | None, list[str]]:
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        values = [_normalize_header(ws.cell(row=r, column=c).value) for c in range(1, scan_cols + 1)]
        # trim trailing empties
        while values and values[-1] == "":
            values.pop()
        if not values:
            continue
        if _looks_like_header_row(values):
            headers = [v for v in values if v]
            return r, headers
    return None, []


def summarize_workbook(excel_path: Path) -> list[SheetSummary]:
    wb = load_workbook(excel_path, data_only=True)
    summaries: list[SheetSummary] = []

    for name in wb.sheetnames:
        ws = wb[name]
        header_row, headers = _find_header_row(ws)
        summaries.append(
            SheetSummary(
                name=name,
                max_row=ws.max_row or 0,
                max_col=ws.max_column or 0,
                header_row=header_row,
                headers=headers,
            )
        )

    return summaries


def main() -> int:
    repo_root = Path(__file__).resolve().parents[2]
    default_excel = repo_root / "הקבצות_מבחן_מיפוי_במקום_הערכה1.xlsx"

    excel_path = default_excel
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return 1

    summaries = summarize_workbook(excel_path)
    print(f"Workbook: {excel_path.name}")
    print(f"Sheets: {len(summaries)}")
    print("")

    for s in summaries:
        print(f"- {s.name}")
        print(f"  size: rows={s.max_row} cols={s.max_col}")
        print(f"  header_row: {s.header_row if s.header_row is not None else 'NOT FOUND'}")
        if s.headers:
            shown = s.headers[:12]
            print(f"  headers: {shown}{' ...' if len(s.headers) > 12 else ''}")
        else:
            print("  headers: []")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
