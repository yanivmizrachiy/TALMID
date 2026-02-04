from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    team_root = Path(__file__).resolve().parents[1]
    repo_root = team_root.parent

    excel_path = repo_root / "הקבצות_מבחן_מיפוי_במקום_הערכה1.xlsx"
    if not excel_path.exists():
        print(f"ERROR: missing excel: {excel_path}")
        return 1

    groups_path = team_root / "נתונים" / "הקבצות.json"
    with groups_path.open("r", encoding="utf-8") as f:
        groups = (json.load(f).get("groups") or [])

    teacher_names = sorted({t for g in groups for t in (g.get("teachers") or []) if t})
    # common variants
    teacher_names += [
        "אוסנת קריפט",
        "אסנת קריפט",
        "אילנית רז",
        "יניב רז",
        "טל נחמיה",
        "רונית פואל",
        "סוניה רפאלי",
        "נעמי שניידר",
        "נורית מויאל",
    ]
    teacher_names = sorted(set(teacher_names))

    wb = load_workbook(excel_path, data_only=True)

    target_sheets = [
        "ז׳ א׳",
        "ז' א1",
        "ח׳ א1",
        "ח׳ א׳1",
        "ט׳ א - 1",
        "ט׳ 2 מקדמת",
    ]

    for sheet in target_sheets:
        if sheet not in wb.sheetnames:
            print(f"- {sheet}: not found")
            continue

        ws = wb[sheet]
        captured: list[str] = []
        for r in range(1, 6):
            for c in range(1, 11):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    captured.append(s)

        joined = " | ".join(captured)
        hits = sorted({t for t in teacher_names if t and t in joined})
        print(f"- {sheet}: teacher_hits={hits if hits else 'NONE'}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
