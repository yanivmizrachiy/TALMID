from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


AUTO_GENERATED_NOTE = "נוצר אוטומטית מייבוא אקסל"


@dataclass(frozen=True)
class GroupFolder:
    grade: str
    subject: str
    group_name: str
    folder: str
    variant: str | None
    class_hint: str | None
    teachers: list[str]


def _read_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _normalize_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ")


def _normalize_homeroom_class(value: str) -> tuple[str, str]:
    """Return (normalized_homeroom, warning).

    We must not confuse homeroom class (כיתת אם) with grouping (הקבצה).
    """

    raw = _normalize_str(value)
    if not raw:
        return "", ""

    # Tokenize while removing quotes/apostrophes
    raw_simple = raw
    for ch in ["'", "׳", '"', "״", "-", ":"]:
        raw_simple = raw_simple.replace(ch, " ")
    tokens = [t for t in raw_simple.split() if t]

    group_tokens = {"א", "א1", "מדעית", "מקדמת"}
    has_group_token = any(t in group_tokens for t in tokens)

    grade_letter = next((t for t in tokens if t in ["ז", "ח", "ט"]), "")
    m = re.search(r"(\d+)", raw_simple)
    digits = m.group(1) if m else ""

    # Example of wrong data: "ח א" or "ט א" => grouping mistakenly placed in class column
    if grade_letter and has_group_token and not digits:
        return "", f"הערה: בעמודת כיתה הופיעה הקבצה ('{raw}')"

    # Normalize: "ז 1" / "ז' 1" / "ז׳ 1" / "ז1" / "ח 2 א" -> "ח2" (with warning)
    if grade_letter and digits:
        warning = ""
        if has_group_token:
            warning = f"הערה: בעמודת כיתה הופיע גם טקסט של הקבצה ('{raw}')"
        return f"{grade_letter}{digits}", warning

    # If cell is just a grouping label
    if len(tokens) == 1 and tokens[0] in group_tokens:
        return "", f"הערה: בעמודת כיתה הופיעה הקבצה ('{raw}')"

    # Unknown format
    return raw, "הערה: ערך כיתה לא זוהה ככיתת אם תקנית"


def _norm_key(text: str) -> str:
    return (
        text.replace("\"", "")
        .replace("'", "")
        .replace("׳", "")
        .replace("״", "")
        .replace(" ", "")
        .replace("-", "")
        .replace("_", "")
        .lower()
    )


def _load_groups(team_root: Path) -> list[GroupFolder]:
    data = _read_json(team_root / "נתונים" / "הקבצות.json")
    groups: list[GroupFolder] = []
    for entry in data.get("groups") or []:
        groups.append(
            GroupFolder(
                grade=entry.get("grade", ""),
                subject=entry.get("subject", ""),
                group_name=entry.get("group_name", ""),
                folder=entry.get("folder", ""),
                variant=entry.get("variant"),
                class_hint=entry.get("class_hint"),
                teachers=list(entry.get("teachers") or []),
            )
        )
    return groups


def _guess_group_folder(sheet_name: str, groups: list[GroupFolder]) -> str | None:
    s = _norm_key(sheet_name)

    candidates: list[tuple[int, GroupFolder]] = []
    for g in groups:
        score = 0
        if g.grade and _norm_key(g.grade) in s:
            score += 3
        if g.group_name and _norm_key(g.group_name) in s:
            score += 3
        if g.class_hint and _norm_key(g.class_hint) in s:
            score += 2
        if g.subject and _norm_key(g.subject) in s:
            score += 1
        if g.variant and _norm_key(g.variant) in s:
            score += 1
        # teacher name in sheet title helps disambiguate (e.g., "נורית מויאל")
        for t in g.teachers:
            if t and _norm_key(t) in s:
                score += 4
        if score > 0:
            candidates.append((score, g))

    if not candidates:
        return None

    candidates.sort(key=lambda t: t[0], reverse=True)
    best_score = candidates[0][0]
    best = [g for score, g in candidates if score == best_score]
    if len(best) != 1:
        return None

    return best[0].folder or None


def _safe_filename(text: str) -> str:
    # Keep Hebrew/letters/digits; replace others with underscore
    cleaned = []
    for ch in text.strip():
        if ch.isalnum() or ch in ["א", "ב", "ג", "ד", "ה", "ו", "ז", "ח", "ט", "י", "כ", "ל", "מ", "נ", "ס", "ע", "פ", "צ", "ק", "ר", "ש", "ת", " ", "׳", "'", "\""]:
            cleaned.append(ch)
        else:
            cleaned.append(" ")
    s = "".join(cleaned)
    # normalize spaces/quotes
    s = s.replace("׳", "").replace("\"", "").replace("'", "")
    s = "_".join([p for p in s.split() if p])
    if not s:
        return "sheet"
    return s[:80]


def _find_header_row(ws, scan_rows: int = 30, scan_cols: int = 30) -> tuple[int | None, list[str]]:
    keywords = ["שם", "משפחה", "פרטי", "תלמיד", "כיתה", "תז", "ת.ז", "תעודת", "זהות"]

    for r in range(1, min(ws.max_row, scan_rows) + 1):
        values = [_normalize_str(ws.cell(row=r, column=c).value) for c in range(1, scan_cols + 1)]
        while values and values[-1] == "":
            values.pop()
        if not values:
            continue

        joined = " ".join(values)
        hits = sum(1 for k in keywords if k in joined)
        non_empty = sum(1 for v in values if v)
        if hits >= 2 and non_empty >= 2:
            return r, values

    return None, []


def _build_column_index(headers: list[str]) -> dict[str, int]:
    return {_normalize_str(h): idx for idx, h in enumerate(headers)}


def _pick_column(headers_map: dict[str, int], aliases: list[str]) -> int | None:
    # exact match
    for a in aliases:
        for h, idx in headers_map.items():
            if _norm_key(h) == _norm_key(a):
                return idx
    # contains match
    for a in aliases:
        ak = _norm_key(a)
        for h, idx in headers_map.items():
            if ak and ak in _norm_key(h):
                return idx
    return None


def _iter_data_rows(ws, start_row: int, max_cols: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for r in range(start_row, ws.max_row + 1):
        values = [_normalize_str(ws.cell(row=r, column=c).value) for c in range(1, max_cols + 1)]
        if not any(v for v in values):
            continue
        rows.append(values)
    return rows


def _split_full_name(full_name: str) -> tuple[str, str]:
    parts = [p for p in full_name.split() if p]
    if len(parts) >= 2:
        return parts[0], parts[-1]
    if len(parts) == 1:
        return parts[0], ""
    return "", ""


def _write_students_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "student_id",
        "first_name",
        "last_name",
        "full_name",
        "homeroom_class",
        "homeroom_class_raw",
        "source_sheet",
        "notes",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def main() -> int:
    team_root = Path(__file__).resolve().parents[1]
    repo_root = team_root.parent

    default_excel = repo_root / "הקבצות_מבחן_מיפוי_במקום_הערכה1.xlsx"
    mapping_path = team_root / "נתונים" / "excel_mapping.json"

    groups = _load_groups(team_root)

    mapping = _read_json(mapping_path) if mapping_path.exists() else {"version": 1, "sheets": {}}
    sheets_mapping: dict[str, dict] = mapping.get("sheets") or {}

    wb = load_workbook(default_excel, data_only=True)

    unmapped: list[str] = []
    summary: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_cfg = sheets_mapping.get(sheet_name) or {}

        if sheet_cfg.get("ignore") is True:
            summary.append({"sheet": sheet_name, "status": "ignored"})
            continue

        folder = sheet_cfg.get("group_folder")

        if not folder:
            folder = _guess_group_folder(sheet_name, groups)

        if not folder:
            unmapped.append(sheet_name)
            summary.append({"sheet": sheet_name, "status": "unmapped"})
            continue

        header_row, headers = _find_header_row(ws)
        if header_row is None:
            unmapped.append(sheet_name)
            summary.append({"sheet": sheet_name, "status": "no_header", "folder": folder})
            continue

        headers_map = _build_column_index(headers)

        columns_cfg = (sheet_cfg.get("columns") or {})
        student_id_aliases = columns_cfg.get("student_id") or ["תז", "ת.ז", "תעודת זהות", "מספר זהות"]
        first_aliases = columns_cfg.get("first_name") or ["שם פרטי"]
        last_aliases = columns_cfg.get("last_name") or ["שם משפחה"]
        full_aliases = columns_cfg.get("full_name") or ["שם מלא", "שם תלמיד", "שם"]
        class_aliases = columns_cfg.get("homeroom_class") or ["כיתה", "כיתת אם"]
        notes_aliases = columns_cfg.get("notes") or ["הערות"]

        idx_id = _pick_column(headers_map, student_id_aliases)
        idx_first = _pick_column(headers_map, first_aliases)
        idx_last = _pick_column(headers_map, last_aliases)
        idx_full = _pick_column(headers_map, full_aliases)
        idx_class = _pick_column(headers_map, class_aliases)
        idx_notes = _pick_column(headers_map, notes_aliases)

        max_cols = max(1, len(headers))
        data_rows = _iter_data_rows(ws, header_row + 1, max_cols)

        students: list[dict[str, str]] = []
        for row in data_rows:
            student_id = row[idx_id] if idx_id is not None and idx_id < len(row) else ""
            first_name = row[idx_first] if idx_first is not None and idx_first < len(row) else ""
            last_name = row[idx_last] if idx_last is not None and idx_last < len(row) else ""
            full_name = row[idx_full] if idx_full is not None and idx_full < len(row) else ""
            homeroom_class_raw = row[idx_class] if idx_class is not None and idx_class < len(row) else ""
            notes = row[idx_notes] if idx_notes is not None and idx_notes < len(row) else ""

            if not full_name and (first_name or last_name):
                full_name = " ".join([p for p in [first_name, last_name] if p])

            if full_name and (not first_name and not last_name):
                first_name, last_name = _split_full_name(full_name)

            # Skip rows that don't look like students
            if not (full_name or first_name or last_name):
                continue

            homeroom_class, warning = _normalize_homeroom_class(homeroom_class_raw)
            if warning:
                notes = (notes + " | " + warning).strip(" |") if notes else warning

            students.append(
                {
                    "student_id": student_id,
                    "first_name": first_name,
                    "last_name": last_name,
                    "full_name": full_name,
                    "homeroom_class": homeroom_class,
                    "homeroom_class_raw": homeroom_class_raw,
                    "source_sheet": sheet_name,
                    "notes": notes,
                }
            )

        output_folder = team_root / folder
        sheet_slug = _safe_filename(sheet_name)
        out_csv = output_folder / f"תלמידים_מהאקסל__{sheet_slug}.csv"
        _write_students_csv(out_csv, students)

        info = {
            "source_excel": str(default_excel.name),
            "source_sheet": sheet_name,
            "group_folder": folder,
            "import_note": AUTO_GENERATED_NOTE,
            "detected": {
                "header_row": header_row,
                "headers": [h for h in headers if h],
                "columns": {
                    "student_id": idx_id,
                    "first_name": idx_first,
                    "last_name": idx_last,
                    "full_name": idx_full,
                    "homeroom_class": idx_class,
                    "notes": idx_notes,
                },
            },
            "counts": {"students": len(students)},
        }
        _write_json(output_folder / f"excel_import_info__{sheet_slug}.json", info)

        summary.append({"sheet": sheet_name, "status": "imported", "folder": folder, "students": len(students)})

        # keep mapping up to date
        sheets_mapping.setdefault(sheet_name, {})
        sheets_mapping[sheet_name]["group_folder"] = folder

    mapping["sheets"] = sheets_mapping
    _write_json(mapping_path, mapping)

    report_lines = [
        "# סיכום יבוא אקסל",
        "",
        f"קובץ מקור: {default_excel.name}",
        "",
    ]
    for item in summary:
        if item["status"] == "imported":
            report_lines.append(f"- {item['sheet']}: יובא ל {item['folder']} ({item['students']} תלמידים)")
        elif item["status"] == "ignored":
            report_lines.append(f"- {item['sheet']}: דולג (ignore)")
        else:
            extra = f" ({item.get('folder')})" if item.get("folder") else ""
            report_lines.append(f"- {item['sheet']}: לא יובא ({item['status']}){extra}")

    if unmapped:
        report_lines.append("")
        report_lines.append("## צריך שייכון ידני")
        report_lines.append("")
        report_lines.append("הגדר לכל אחד מהגיליונות הבאים `group_folder` בתוך `נתונים/excel_mapping.json`:" )
        for name in unmapped:
            report_lines.append(f"- {name}")

    (team_root / "דוחות").mkdir(parents=True, exist_ok=True)
    (team_root / "דוחות" / "סיכום_יבוא_אקסל.md").write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    # Print only sheet names + statuses (no student names)
    imported = sum(1 for s in summary if s["status"] == "imported")
    print(f"Imported sheets: {imported}/{len(summary)}")
    if unmapped:
        print("Unmapped sheets (need folder mapping):")
        for name in unmapped:
            print(f"- {name}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
